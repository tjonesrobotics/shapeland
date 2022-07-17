import openpyxl as pyxl
import datetime
expectedSheets = ["Metadata", "Park", "Attractions",\
	"Activities", "Restaurants", "Waypoints", \
	"WalkingRoutes"]

def readSheetNames(wb):
	#Read the sheet names
	sheets = [s for s in wb.sheetnames if s in expectedSheets]

	#If there's an extra sheet, warn
	sheetWarningFlag=False
	for s in wb.sheetnames:
		if s not in sheets:
			if not sheetWarningFlag:
				print("********************************************************")
				print("********************************************************")
			print("  WARNING: UNEXPECTED SHEET PRESENT\t", s)
			sheetWarningFlag=True
	#If there's a missing sheet, warn
	for s in expectedSheets:
		if s not in sheets:
			if not sheetWarningFlag:
				print("********************************************************")
				print("********************************************************")
			print("  WARNING: EXPECTED SHEET NOT PRESENT\t", s)
			sheetWarningFlag=True

	#If there was a sheet warning, display the block end
	if sheetWarningFlag:
		print("********************************************************")
		print("********************************************************\n")
	return sheets

def readSheetRow(sheet,row):
	row=str(row)
	i=1
	data=[]
	while True:
		c=pyxl.utils.cell.get_column_letter(i)
		next=sheet[c+row].value

		#If you get a blank, verify if the next one is blank
		if next==None:
			i+=1
			c=pyxl.utils.cell.get_column_letter(i)
			next=sheet[c+row].value
			if next==None:
				break

		#Append the value to the data list
		data.append(next)
		i+=1

	#If nothing was read, then return none
	if data==[]: return None

	#If nothing valid was read, return an empty list
	if data[0]=="": data=[]
	if data[0][0]=="#": data=[]

	return data

def dataProcessError(message,data):
	print("********************************************************")
	print(message)
	print(data)
	print("********************************************************")
	exit()


def readVarsFromSheet(ws):
	variables=[]
	data=[""]
	row=1
	while True:
		data=readSheetRow(ws,row)

		#Increment the row. If there's no data, continue or break
		row+=1
		if data==None: break
		if len(data)==0: continue

		if len(data)==1:
			dataProcessError("Not a key data pair on Sheet "+ws.title,data)
		if len(data)>2 and len(data)%2!=1:
			dataProcessError("Not a key/data number mismatch on Sheet "+ws.title,data)

		if data[1]=="=TRUE()": data[1]=True
		if data[1]=="=FALSE()": data[1]=False

		#If there are more than one entry, turn it into a dict
		if len(data)>2:
			newDict = {}
			for i in range(len(data)//2):
				key=data[1+2*i]
				if type(key)==datetime.time:
					key = key.strftime("%I:%M %p")
				val=data[2+2*i]
				if val=="=TRUE()": val=True
				if val=="=FALSE()": val=False
				newDict[key]=val
			data=[data[0],newDict]

		variables.append(data)
	return variables


def readEntitiesFromSheet(ws):
	data=[""]
	entities=[]
	row=1
	headers=None
	while True:
		data=readSheetRow(ws,row)

		#Increment the row. If there's no data, continue or break
		row+=1
		if data==None: break
		if len(data)==0: continue

		if headers==None:
			headers=data
			continue

		newDict={}
		if len(headers)!=len(data):
			dataProcessError("Cannot process data entry. Column number mismatch on Sheet "+ws.title,data)
		for h,d in zip(headers,data):
			if d=="=TRUE()": d=True
			if d=="=FALSE()": d=False
			newDict[h]=d

		entities.append(newDict)
	return entities

def readSimParams(fileName):
	wb = pyxl.load_workbook(fileName)
	sheets = readSheetNames(wb)


	#Read each sheet, based on the type of data it stores
	variables = []
	attractions, activities, restaurants, waypoints = [], [], [], []
	for sName in sheets:
		ws=wb[sName]
		if sName in ["Metadata","Park"]:
			variables+=readVarsFromSheet(ws)
		elif sName in ["Attractions"]:
			attractions=readEntitiesFromSheet(ws)
		elif sName in ["Activities"]:
			activities=readEntitiesFromSheet(ws)
		elif sName in ["Restaurants"]:
			restaurants=readEntitiesFromSheet(ws)
		elif sName in ["Waypoints"]:
			waypoints=readEntitiesFromSheet(ws)
		elif sName in ["WalkingRoutes"]:
			print("\n!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
			print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
			print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
			print("   FUNCTION NOT DEFINED - Walk Times")
			print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
			print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
			print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!\n")
			'''TODO'''

	return variables, attractions, activities, restaurants, waypoints

